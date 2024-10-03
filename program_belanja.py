from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime 
import os
import sys
import time
import requests
import random
from colorama import Fore

os.system("clear")

waktu = time.strftime("%d-%m-%Y", time.localtime())
jam = time.strftime("%H:%M", time.localtime())
ip = requests.get('https://api.ipify.org').text

B = Fore.BLUE
W = Fore.WHITE
R = Fore.RED
G = Fore.GREEN
BL = Fore.BLACK
Y = Fore.YELLOW

hijau = "\033[0;92m"
putih = "\033[0;97m"
abu = "\033[0;90m"
kuning = "\033[0;93m"
ungu = "\033[0;95m"
merah = "\033[0;91m"
biru = "\033[0;96m"

def clear_screen():
    print("\033[H\033[J", end="")

clear_screen()

def autoketik(s):
    for c in s + "\n":
        sys.stdout.write(c)
        sys.stdout.flush()
        time.sleep(0.05)

def banner():
    print(f"""=============================================\n{putih}[{B}•{putih}] {biru}Developer{putih} : {hijau}MarshallSetian
{putih}[{B}•{putih}] {ungu}Instagram {putih}: @marshall_setian
{W}[{B}•{W}]{putih} Ip Kamu {putih}  :{Y} {ip}
{W}[{B}•{W}] {putih}Waktu     {putih}:{merah} {waktu}
{W}[{B}•{W}] {putih}Jam       {putih}:{biru} {jam}
{putih}===========================
{hijau}PROGRAM KASIR CALON JURAGAN
{putih}===========================
Bahasa Program :{hijau} Python{putih}
Fitur : {kuning}Mode Kasir,Save Excell,View-LocalHost{putih}
=============================================""")

banner()
    


def jumlah_barang():
    total_belanja = 0
    daftar_barang = []  # List untuk menyimpan data barang
    while True:
        print("\nMasukkan Nama Barang Belanja\n")
        
        nama_barang = input("Nama Barang           : ")
        
        if nama_barang.lower() == '0':
            break
        
        jumlah = int(input("Jumlah Barang         : "))
        harga = float(input("Harga Barang per unit : "))
        
        # Menghitung total belanja
        total_harga = jumlah * harga
        total_belanja += total_harga
        print(f"\nTotal Belanja Untuk {jumlah} {nama_barang} = {kuning}Rp. {total_harga:,}{putih}")
        
        # Menyimpan data barang ke dalam daftar
        daftar_barang.append([nama_barang, jumlah, harga, total_harga])

        # Menambahkan opsi untuk menambah barang
        print(f"\nKetik {hijau}'1'{putih} ==> Menambahkan Barang.")
        print(f"Ketik {hijau}'0'{putih} ==> Selesai.")
        tambah_barang = input("\nMasukkan Pilihan Dengan Angka : ")
        if tambah_barang.lower() == '0':
            break

    return total_belanja, daftar_barang

def warung_belanja(uang_awal, total_belanja):
    sisa_uang = uang_awal - total_belanja
    return sisa_uang

# Memanggil fungsi untuk mendapatkan total belanja
total_belanja, daftar_barang = jumlah_barang()

# Menampilkan total belanja terlebih dahulu
autoketik(f"\nTotal Belanja: {kuning}Rp. {total_belanja:,.2f}{putih}")

# Meminta input jumlah uang yang dibawa setelah total belanja ditampilkan
uang_awal = float(input("\nMasukkan Jumlah Uang Pelanggan : Rp. "))

# Menghitung sisa uang
sisa_uang = warung_belanja(uang_awal, total_belanja)

# Menampilkan informasi akhir
print(f"\nUang Pelanggan : {kuning}Rp. {uang_awal:,.2f}{putih}")
print(f"Total Belanja  : {kuning}Rp. {total_belanja:,.2f}{putih}")
autoketik(f"\nKembalian      : {biru}Rp. {sisa_uang:,.2f}{putih}")

# Opsi untuk menyimpan data ke Excel
save_option = input(f"\nKetik {hijau}'99'{putih} untuk menyimpan data ke dalam file Excel: ")
if save_option == '99':
    filename = "data_belanja.xlsx"  # Nama file Excel
    
    # Mendapatkan tanggal dan waktu saat ini
    tanggal_sekarang = datetime.now().strftime("%d-%m-%Y")
    waktu_sekarang = datetime.now().strftime("%H:%M")
    
    # Membaca workbook yang ada
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # Menambahkan header jika file baru
        ws.append(["Tanggal", "Jam", "Nama Barang", "Jumlah", "Harga per Unit", "Total Harga"])

    # Menambahkan data baru dari daftar_barang ke worksheet
    for row in daftar_barang:
        # Menambahkan tanggal dan waktu di kolom yang berbeda
        ws.append([tanggal_sekarang, waktu_sekarang] + row)

    # Mengatur lebar kolom agar rapi
    ws.column_dimensions['A'].width = 12  # Tanggal
    ws.column_dimensions['B'].width = 8   # Jam
    ws.column_dimensions['C'].width = 25  # Nama Barang
    ws.column_dimensions['D'].width = 8   # Jumlah
    ws.column_dimensions['E'].width = 15  # Harga per Unit
    ws.column_dimensions['F'].width = 15  # Total Harga
    
    # Format seluruh data di Excel untuk membuatnya rata tengah dan rapi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Rata tengah
            cell.border = thin_border  # Tambahkan border tipis
            if cell.column == 5 or cell.column == 6:  # Format kolom Harga per Unit dan Total Harga
                cell.number_format = 'Rp #,##0.00'  # Format harga dengan dua desimal

    # Menyimpan workbook kembali
    wb.save(filename)
    autoketik(f"{hijau}Data Belanja Telah Ditambahkan ke File '{filename}'{putih}")

